#!/usr/bin/env python3
"""
Aktualizace cenových dat z REE_Price_Analytics.xlsx.

Použití:
  python3 scripts/update_prices.py [cesta_k_excelu]

Výchozí cesta: ~/Downloads/REE_Price_Analytics.xlsx
Výstup: public/data/prices_history.json (+ summary_stats, correlation, events, volatility)
"""

import json
import os
import sys
from datetime import datetime

import openpyxl

# --- Config ---
DEFAULT_XLSX = os.path.expanduser("~/Downloads/REE_Price_Analytics.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "public", "data")


def to_json_value(val):
    """Convert Excel cell value to JSON-safe value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        return round(val, 4) if isinstance(val, float) else val
    return str(val)


def read_sheet(wb, sheet_name, header_row=1):
    """Read sheet into list of dicts."""
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h is not None]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = to_json_value(ws.cell(r, c).value)
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def save_json(data, filename):
    """Save data to JSON file."""
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {filename} ({len(data) if isinstance(data, list) else 'object'})")


def extract_correlation(wb):
    """Extract correlation matrix from Correlation sheet."""
    ws = wb["Correlation"]
    # Find header row (first row with a value in column B)
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value is not None:
            header_row = r
            break
    if header_row is None:
        return {"labels": [], "matrix": []}

    labels = []
    for c in range(2, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            labels.append(str(val))

    matrix = []
    for r in range(header_row + 1, header_row + 1 + len(labels)):
        row = []
        for c in range(2, 2 + len(labels)):
            val = ws.cell(r, c).value
            row.append(round(val, 4) if isinstance(val, (int, float)) else None)
        matrix.append(row)

    return {"labels": labels, "matrix": matrix}


def extract_summary_stats(wb):
    """Extract summary stats grouped by period."""
    ws = wb["Summary_Stats"]
    periods = []
    period_starts = [2, 15, 28, 41, 54]

    for start_row in period_starts:
        period_label = ws.cell(start_row, 1).value
        if not period_label:
            continue

        headers = [ws.cell(start_row + 1, c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h is not None]

        stats = []
        for r in range(start_row + 2, start_row + 12):
            row = {}
            for c, h in enumerate(headers, 1):
                val = ws.cell(r, c).value
                row[h] = round(val, 4) if isinstance(val, (int, float)) else to_json_value(val)
            if row.get(headers[0]):
                stats.append(row)

        periods.append({"period": str(period_label), "stats": stats})

    return {"periods": periods}


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

    if not os.path.exists(xlsx_path):
        print(f"Soubor nenalezen: {xlsx_path}")
        sys.exit(1)

    print(f"Načítám: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 1. Monthly prices
    prices = read_sheet(wb, "Monthly_USD")
    # Rename 'Date' field for consistency
    for row in prices:
        if "Date" not in row and "date" in row:
            row["Date"] = row.pop("date")
    save_json(prices, "prices_history.json")

    # 2. Weekly prices
    weekly = read_sheet(wb, "Weekly_USD")
    for row in weekly:
        if "Date" not in row and "date" in row:
            row["Date"] = row.pop("date")
    save_json(weekly, "prices_weekly.json")

    # 3. Events
    events = read_sheet(wb, "Events")
    save_json(events, "events.json")

    # 4. Correlation matrix
    corr = extract_correlation(wb)
    save_json(corr, "correlation.json")

    # 5. Summary stats
    stats = extract_summary_stats(wb)
    save_json(stats, "summary_stats.json")

    # 6. Volatility
    vol = read_sheet(wb, "Volatility_12M")
    save_json(vol, "volatility.json")

    # Summary
    last_date = prices[-1].get("Date", "?") if prices else "?"
    print(f"\nHotovo! Poslední datum v cenových datech: {last_date}")
    print(f"Celkem {len(prices)} měsíčních záznamů.")


if __name__ == "__main__":
    main()
